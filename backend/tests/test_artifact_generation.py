import tempfile
import unittest
from pathlib import Path
import sys
import gc
import shutil
from uuid import uuid4

import pandas as pd
from openpyxl import load_workbook
from pypdf import PdfReader

BASE_DIR = Path(__file__).resolve().parents[1]
if str(BASE_DIR) not in sys.path:
    sys.path.insert(0, str(BASE_DIR))

from schemas.artifact_models import UploadedFileInfo
from services.artifact_generation_service import generate_artifact, list_all_artifacts
from services.multi_file_normalization_service import normalize_multiple_excel_files


class ArtifactGenerationTests(unittest.TestCase):
    def test_list_all_artifacts_includes_chart_and_analysis(self):
        items = list_all_artifacts()
        artifact_ids = {(item.artifact_type, item.artifact_id) for item in items}
        self.assertIn(("chart", "income_expense_pie_chart"), artifact_ids)
        self.assertIn(("analysis", "financial_risk_analysis"), artifact_ids)
        self.assertIn(("report", "income_expense_report"), artifact_ids)

    def test_prompt_only_chart_generation_creates_jpg(self):
        audit_run_id = "test_prompt_only_chart_generation"
        result = generate_artifact(
            artifact_type="chart",
            artifact_id="income_expense_pie_chart",
            audit_run_id=f"{audit_run_id}_{uuid4().hex[:8]}",
            user_prompt="Ocak gelir 500 bin TL, gider 300 bin TL. Subat gelir 700 bin TL, gider 400 bin TL.",
        )
        self.assertIn(result.status, {"success", "warning"})
        self.assertEqual(result.output_format, "jpg")
        self.assertTrue(result.output_file_path)
        self.assertTrue(Path(result.output_file_path).exists())

    def test_excel_report_generation_uses_standard_sheets_and_charts(self):
        sample_file = BASE_DIR / "storage/uploads/2355ef0ba2594b63829f5f5b9d438cb5/gelir_gider_6_ay_200_satir_test_verisi.xlsx"
        if not sample_file.exists():
            self.skipTest("Sample Excel dosyasi bulunamadi.")
        result = generate_artifact(
            artifact_type="report",
            artifact_id="income_expense_report",
            audit_run_id=f"test_standard_excel_report_{uuid4().hex[:8]}",
            file_paths=[str(sample_file)],
            user_prompt="son 3 ay ve 50 bin uzeri islemler",
        )
        self.assertIn(result.status, {"success", "warning"})
        workbook = load_workbook(result.output_file_path)
        for sheet in ["Rapor", "Grafikler", "Normalize Veri", "Metodoloji ve Ek Bilgi"]:
            self.assertIn(sheet, workbook.sheetnames)
        charts_sheet = workbook["Grafikler"]
        self.assertGreaterEqual(len(getattr(charts_sheet, "_charts", [])), 1)
        self.assertLessEqual(len(getattr(charts_sheet, "_charts", [])), 4)
        workbook.close()

    def test_excel_analysis_generation_creates_a4_pdf_with_minimum_pages(self):
        sample_file = BASE_DIR / "storage/uploads/2355ef0ba2594b63829f5f5b9d438cb5/gelir_gider_6_ay_200_satir_test_verisi.xlsx"
        if not sample_file.exists():
            self.skipTest("Sample Excel dosyasi bulunamadi.")
        result = generate_artifact(
            artifact_type="analysis",
            artifact_id="financial_risk_analysis",
            audit_run_id=f"test_excel_analysis_generation_{uuid4().hex[:8]}",
            file_paths=[str(sample_file)],
            user_prompt="son 3 ay ve 50 bin uzeri islemler",
        )
        self.assertIn(result.status, {"success", "warning"})
        self.assertEqual(result.output_format, "pdf")
        self.assertTrue(result.output_file_path)
        self.assertTrue(Path(result.output_file_path).exists())

        with open(result.output_file_path, "rb") as pdf_stream:
            reader = PdfReader(pdf_stream)
            self.assertGreaterEqual(len(reader.pages), 5)
            first_page = reader.pages[0]
            self.assertAlmostEqual(float(first_page.mediabox.width), 595.2756, places=1)
            self.assertAlmostEqual(float(first_page.mediabox.height), 841.8898, places=1)
            extracted_text = "\n".join(page.extract_text() or "" for page in reader.pages[:3])
            self.assertIn("Kapak ve Yonetici Ozeti", extracted_text)
            self.assertIn("Temel Finansal Metrikler", extracted_text)
            self.assertIn("Risk Faktorleri", "\n".join(page.extract_text() or "" for page in reader.pages))

    def test_multi_file_normalization_merges_and_filters(self):
        temp_path = Path(tempfile.mkdtemp())
        try:
            income_file = temp_path / "income.xlsx"
            expense_file = temp_path / "expense.xlsx"
            pd.DataFrame(
                [
                    {"Tarih": "2026-01-15", "Tutar": 150000, "Yon": "income", "Aciklama": "Satis"},
                    {"Tarih": "2026-02-10", "Tutar": 180000, "Yon": "income", "Aciklama": "Tahsilat"},
                ]
            ).to_excel(income_file, index=False)
            pd.DataFrame(
                [
                    {"Islem Tarihi": "2026-01-18", "Islem Tutari": 90000, "Hareket Tipi": "expense", "Not": "Kira"},
                    {"Islem Tarihi": "2026-02-12", "Islem Tutari": 120000, "Hareket Tipi": "expense", "Not": "Maas"},
                ]
            ).to_excel(expense_file, index=False)

            result = normalize_multiple_excel_files(
                files=[
                    UploadedFileInfo(file_name=income_file.name, file_path=str(income_file)),
                    UploadedFileInfo(file_name=expense_file.name, file_path=str(expense_file)),
                ],
                artifact_id="income_expense_report",
                artifact_type="report",
                user_prompt="2026 ve 100 bin uzeri islemler",
                audit_run_id="test_multi_file_normalization",
            )

            self.assertIn(result.status, {"success", "warning"})
            self.assertEqual(len(result.file_summaries), 2)
            self.assertGreaterEqual(result.normalized_row_count, 2)
            self.assertTrue(any(record.get("direction") == "income" for record in result.normalized_records))
            self.assertTrue(any(record.get("direction") == "expense" for record in result.normalized_records))
        finally:
            gc.collect()
            shutil.rmtree(temp_path, ignore_errors=True)

    def test_multi_file_duplicate_warning_and_partial_success(self):
        temp_path = Path(tempfile.mkdtemp())
        try:
            valid_file = temp_path / "valid.xlsx"
            duplicate_file = temp_path / "duplicate.xlsx"
            invalid_file = temp_path / "invalid.csv"

            rows = [
                {"Tarih": "2026-03-01", "Tutar": 75000, "Yon": "expense", "Aciklama": "Lojistik"},
                {"Tarih": "2026-03-05", "Tutar": 92000, "Yon": "income", "Aciklama": "Tahsilat"},
            ]
            pd.DataFrame(rows).to_excel(valid_file, index=False)
            pd.DataFrame(rows).to_excel(duplicate_file, index=False)
            invalid_file.write_text("bad,file\n1,2\n", encoding="utf-8")

            result = normalize_multiple_excel_files(
                files=[
                    UploadedFileInfo(file_name=valid_file.name, file_path=str(valid_file)),
                    UploadedFileInfo(file_name=duplicate_file.name, file_path=str(duplicate_file)),
                    UploadedFileInfo(file_name=invalid_file.name, file_path=str(invalid_file)),
                ],
                artifact_id="income_expense_report",
                artifact_type="report",
                audit_run_id="test_multi_file_duplicate_warning",
            )

            self.assertEqual(result.status, "warning")
            self.assertLess(result.normalized_row_count, result.input_row_count)
            warning_messages = [item.get("message", "") for item in result.warnings]
            self.assertTrue(any("tekrar eden" in message for message in warning_messages))
            self.assertTrue(any(summary.status == "failed" for summary in result.file_summaries))
        finally:
            gc.collect()
            shutil.rmtree(temp_path, ignore_errors=True)


if __name__ == "__main__":
    unittest.main()
