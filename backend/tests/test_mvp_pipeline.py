import json
from datetime import date
from pathlib import Path
import shutil
import unittest
from unittest.mock import patch

from openpyxl import load_workbook
import pandas as pd

from agents.prompt_parsing_agent import extract_prompt_transactions
from agents.report_generator_agent import generate_report
from config import settings
from outputs.output_generator import generate_outputs
from schemas.report_filters import CategoryFilter, ReportFilterSpec
from services.report.report_filter_engine import apply_report_filters
from services.report.report_generation_service import append_filter_sheet, coerce_error_items
from services.report.report_prompt_filter_service import extract_report_filters
from services.report.report_execution_service import ReportExecutionService
from services.report.report_registry_service import ReportRegistryService
from utils.mapping_utils import column_mapping, empty_mapping, fields_for_report, match_field_by_alias
from validators.transaction_validator import date_in_range
from validators.transaction_validator import validate_transactions


class ReportPlatformTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.registry = ReportRegistryService()
        cls.llm_patch = patch("services.report_prompt_filter_service.try_extract_with_llm", return_value=None)
        cls.llm_patch.start()

    @classmethod
    def tearDownClass(cls):
        cls.llm_patch.stop()

    def test_registry_loads_all_new_reports(self):
        report_ids = [definition["report_id"] for definition in self.registry.list_reports()]
        self.assertEqual(
            report_ids,
            [
                "income_expense_report",
                "cash_flow_report",
                "debt_receivable_report",
                "vat_summary_report",
                "personnel_expense_report",
                "sales_performance_report",
                "profitability_report",
                "current_account_report",
                "payroll_cost_report",
                "inventory_cost_report",
                "tax_calculation_report",
            ],
        )

    def test_personnel_expense_report_calculates_total_employer_cost(self):
        report_definition = self.registry.get_report_definition("personnel_expense_report")
        validation = validate_transactions(sample_personnel_expense_records(), report_definition=report_definition)

        self.assertTrue(validation["valid"], validation["errors"])
        self.assertTrue(any(w["field"] == "employer_cost" and w["type"] == "derived_value" for w in validation["warnings"]))
        self.assertEqual(float(validation["transactions"][0]["employer_cost"]), 8325.0)
        self.assertEqual(float(validation["transactions"][0]["total_employer_cost"]), 51325.0)

        report_result = generate_report(
            report_definition=report_definition,
            report_input=validation["transactions"],
            audit_context=validation["audit_context"],
            input_warnings=validation["warnings"],
        )
        self.assertEqual(report_result["summary"]["total_employer_cost"], 91575.0)
        self.assertGreater(report_result["summary"]["warning_count"], 0)

    def test_personnel_expense_report_keeps_total_employer_cost_separate_from_employer_cost(self):
        report_definition = self.registry.get_report_definition("personnel_expense_report")
        validation = validate_transactions(sample_personnel_expense_record_with_input_total_cost(), report_definition=report_definition)

        self.assertTrue(validation["valid"], validation["errors"])
        transaction = validation["transactions"][0]
        self.assertEqual(float(transaction["employer_cost"]), 8325.0)
        self.assertEqual(float(transaction["total_employer_cost"]), 51325.0)
        self.assertFalse(any(w["field"] == "total_employer_cost" and w["type"] == "mismatch" for w in validation["warnings"]))

    def test_personnel_mapping_aliases_distinguish_sgk_and_total_employer_cost(self):
        report_definition = self.registry.get_report_definition("personnel_expense_report")
        candidate_fields = report_definition["input_contract"]["required_fields"] + report_definition["input_contract"]["optional_fields"]

        self.assertEqual(match_field_by_alias("SGK İşveren Payı", candidate_fields), "employer_cost")
        self.assertEqual(match_field_by_alias("İşveren Toplam Maliyeti", candidate_fields), "total_employer_cost")

    def test_payroll_cost_report_derives_sgk_and_total(self):
        report_definition = self.registry.get_report_definition("payroll_cost_report")
        validation = validate_transactions(sample_payroll_records(), report_definition=report_definition)

        self.assertTrue(validation["valid"], validation["errors"])
        self.assertTrue(any(w["field"] == "sgk_employer" for w in validation["warnings"]))
        self.assertEqual(float(validation["transactions"][0]["sgk_employer"]), 13500.0)
        self.assertEqual(float(validation["transactions"][0]["total_employer_cost"]), 77500.0)

    def test_vat_report_normalizes_decimal_tax_rate(self):
        report_definition = self.registry.get_report_definition("vat_summary_report")
        validation = validate_transactions(sample_vat_records_decimal_rate(), report_definition=report_definition)

        self.assertTrue(validation["valid"], validation["errors"])
        self.assertTrue(any(w["field"] == "tax_amount" and w["type"] == "derived_value" for w in validation["warnings"]))
        self.assertEqual(float(validation["transactions"][0]["tax_rate"]), 20.0)
        self.assertEqual(float(validation["transactions"][0]["tax_amount"]), 20000.0)

        report_result = generate_report(
            report_definition=report_definition,
            report_input=validation["transactions"],
            audit_context=validation["audit_context"],
            input_warnings=validation["warnings"],
        )
        self.assertEqual(report_result["summary"]["calculated_vat"], 20000.0)
        self.assertEqual(report_result["summary"]["deductible_vat"], 4000.0)
        self.assertEqual(report_result["summary"]["net_vat"], 16000.0)

    def test_tax_report_rejects_invalid_tax_rate(self):
        report_definition = self.registry.get_report_definition("tax_calculation_report")
        validation = validate_transactions(sample_invalid_tax_rate_records(), report_definition=report_definition)

        self.assertTrue(validation["valid"])
        self.assertEqual(validation["execution_status"], "warning")
        self.assertEqual(validation["usable_row_count"], 0)
        self.assertTrue(any(w["type"] == "invalid_tax_rate" and w["severity"] == "critical" for w in validation["warnings"]))

    def test_tax_report_derives_periodic_summary(self):
        report_definition = self.registry.get_report_definition("tax_calculation_report")
        validation = validate_transactions(sample_tax_records_decimal_rate(), report_definition=report_definition)

        self.assertTrue(validation["valid"], validation["errors"])
        self.assertTrue(any(w["field"] == "tax_amount" and w["type"] == "derived_value" for w in validation["warnings"]))

        report_result = generate_report(
            report_definition=report_definition,
            report_input=validation["transactions"],
            audit_context=validation["audit_context"],
            input_warnings=validation["warnings"],
        )
        periodic_rows = report_result["tables"]["periodic"]
        self.assertEqual(periodic_rows[0]["KDV"], 20000.0)
        self.assertEqual(periodic_rows[0]["Stopaj"], 4000.0)
        self.assertEqual(periodic_rows[0]["Toplam Vergi"], 24000.0)

    def test_sales_report_supports_mismatch_and_refund(self):
        report_definition = self.registry.get_report_definition("sales_performance_report")
        validation = validate_transactions(sample_sales_records_with_refund(), report_definition=report_definition)

        self.assertTrue(validation["valid"], validation["errors"])
        self.assertTrue(any(w["type"] == "recalculated_total_sales" and w["field"] == "total_sales" for w in validation["warnings"]))
        self.assertTrue(any(w["type"] == "refund_detected" for w in validation["warnings"]))

        report_result = generate_report(
            report_definition=report_definition,
            report_input=validation["transactions"],
            audit_context=validation["audit_context"],
            input_warnings=validation["warnings"],
        )
        self.assertEqual(report_result["summary"]["gross_sales"], 95000.0)
        self.assertEqual(report_result["summary"]["refund_total"], 50000.0)
        self.assertEqual(report_result["summary"]["net_sales"], 45000.0)
        self.assertEqual(report_result["summary"]["gross_quantity"], 2.0)
        self.assertEqual(report_result["summary"]["refund_quantity"], 1.0)
        self.assertEqual(report_result["summary"]["net_quantity"], 1.0)
        self.assertEqual(report_result["summary"]["gross_order_count"], 1)
        self.assertEqual(report_result["summary"]["refund_order_count"], 1)
        self.assertEqual(report_result["summary"]["net_order_count"], 0)
        self.assertEqual(report_result["summary"]["net_average_order_value"], 45000.0)
        summary_row = report_result["tables"]["summary"][0]
        self.assertEqual(summary_row["En Cok Satan Urun (Ciro)"], "Danismanlik Hizmeti")
        self.assertEqual(summary_row["En Cok Satan Urun (Adet)"], "Danismanlik Hizmeti")

    def test_sales_report_drops_unknown_transaction_type(self):
        report_definition = self.registry.get_report_definition("sales_performance_report")
        validation = validate_transactions(sample_sales_records_with_unknown_type(), report_definition=report_definition)

        self.assertTrue(validation["valid"], validation["errors"])
        self.assertEqual(len(validation["transactions"]), 1)
        unknown_warning = next(w for w in validation["warnings"] if w["type"] == "unknown_transaction_type")
        self.assertEqual(unknown_warning["action"], "row_dropped")
        self.assertEqual(unknown_warning["value"], "weird")

    def test_sales_report_defaults_to_sale_when_transaction_type_absent_and_return_status_exists(self):
        report_definition = self.registry.get_report_definition("sales_performance_report")
        validation = validate_transactions(sample_sales_records_with_return_status_only(), report_definition=report_definition)

        self.assertTrue(validation["valid"], validation["errors"])
        self.assertEqual(validation["usable_row_count"], 2)
        self.assertFalse(any(w["type"] == "unknown_transaction_type" for w in validation["warnings"]))
        self.assertTrue(any(w["type"] == "return_status_detected" for w in validation["warnings"]))
        self.assertEqual(validation["transactions"][0]["transaction_type"], "sale")
        self.assertEqual(validation["transactions"][0]["return_status"], "partial_return")

        report_result = generate_report(
            report_definition=report_definition,
            report_input=validation["transactions"],
            audit_context=validation["audit_context"],
            input_warnings=validation["warnings"],
        )
        self.assertEqual(report_result["summary"]["gross_sales"], 2000.0)
        self.assertEqual(report_result["summary"]["refund_total"], 0.0)
        self.assertEqual(report_result["summary"]["net_sales"], 2000.0)

    def test_sales_mapping_aliases_map_return_status_not_transaction_type(self):
        report_definition = self.registry.get_report_definition("sales_performance_report")
        candidate_fields = report_definition["input_contract"]["required_fields"] + report_definition["input_contract"]["optional_fields"]

        self.assertEqual(match_field_by_alias("İade Durumu", candidate_fields), "return_status")
        self.assertEqual(match_field_by_alias("Return Status", candidate_fields), "return_status")
        self.assertEqual(match_field_by_alias("Refund Status", candidate_fields), "return_status")

    def test_sales_mapping_repairs_return_status_column_misclassified_as_transaction_type(self):
        report_definition = self.registry.get_report_definition("sales_performance_report")
        execution = ReportExecutionService(self.registry)
        field_mappings = {field_name: empty_mapping(field_name) for field_name in fields_for_report(report_definition)}
        field_mappings["date"] = column_mapping("date", "Tarih")
        field_mappings["product_name"] = column_mapping("product_name", "Urun")
        field_mappings["customer"] = column_mapping("customer", "Musteri")
        field_mappings["quantity"] = column_mapping("quantity", "Miktar")
        field_mappings["unit_price"] = column_mapping("unit_price", "Birim Fiyat")
        field_mappings["transaction_type"] = column_mapping("transaction_type", "Return Status")

        mapping_json = {
            "status": "passed",
            "report_type": "sales_performance_report",
            "selected_sheet": "Sheet1",
            "confidence": 0.9,
            "missing_fields": [],
            "field_mappings": field_mappings,
            "warnings": [],
            "message": None,
        }

        normalized = execution.normalize_for_report(
            report_definition=report_definition,
            raw_data=sample_sales_excel_rows_with_return_status_column(),
            mapping_json=mapping_json,
            intent={},
        )
        validation = validate_transactions(normalized["items"], report_definition=report_definition)

        self.assertTrue(validation["valid"], validation["errors"])
        self.assertEqual(validation["usable_row_count"], 2)
        self.assertFalse(any(w["type"] == "unknown_transaction_type" for w in validation["warnings"]))
        self.assertTrue(any(w["type"] == "return_status_detected" for w in validation["warnings"]))
        self.assertEqual(validation["transactions"][0]["transaction_type"], "sale")
        self.assertEqual(validation["transactions"][0]["return_status"], "partial_return")
        self.assertEqual(mapping_json["field_mappings"]["transaction_type"]["mapping_type"], "not_available")
        self.assertEqual(mapping_json["field_mappings"]["return_status"]["source_column"], "Return Status")

    def test_sales_report_all_unknown_rows_leave_no_usable_data(self):
        report_definition = self.registry.get_report_definition("sales_performance_report")
        validation = validate_transactions(sample_sales_records_all_unknown_type(), report_definition=report_definition)

        self.assertTrue(validation["valid"], validation["errors"])
        self.assertEqual(validation["usable_row_count"], 0)
        self.assertTrue(all(w["type"] == "unknown_transaction_type" for w in validation["warnings"]))
        with self.assertRaisesRegex(ValueError, "Validation sonrasi rapor uretecek veri kalmadi."):
            generate_report(
                report_definition=report_definition,
                report_input=validation["transactions"],
                audit_context=validation["audit_context"],
                input_warnings=validation["warnings"],
            )

    def test_sales_report_warns_for_negative_normal_sale(self):
        report_definition = self.registry.get_report_definition("sales_performance_report")
        validation = validate_transactions(sample_negative_sale_records(), report_definition=report_definition)

        self.assertTrue(validation["valid"], validation["errors"])
        negative_warning = next(w for w in validation["warnings"] if w["type"] == "negative_sale_total")
        self.assertEqual(negative_warning["severity"], "warning")
        self.assertEqual(negative_warning["action"], "row_retained")

    def test_sales_leaderboards_use_net_sales_after_refunds(self):
        report_definition = self.registry.get_report_definition("sales_performance_report")
        validation = validate_transactions(sample_sales_leaderboard_records(), report_definition=report_definition)
        report_result = generate_report(
            report_definition=report_definition,
            report_input=validation["transactions"],
            audit_context=validation["audit_context"],
            input_warnings=validation["warnings"],
        )

        self.assertEqual(report_result["summary"]["top_customer"], "Beta AS")
        self.assertEqual(report_result["summary"]["top_salesperson"], "Selin")
        self.assertEqual(report_result["summary"]["top_product_by_revenue"], "Urun B")
        self.assertEqual(report_result["summary"]["top_product_by_quantity"], "Urun B")

    def test_inventory_report_handles_negative_stock_and_missing_stock_out_cost(self):
        report_definition = self.registry.get_report_definition("inventory_cost_report")
        validation = validate_transactions(sample_inventory_negative_stock_records(), report_definition=report_definition)

        self.assertTrue(validation["valid"], validation["errors"])
        report_result = generate_report(
            report_definition=report_definition,
            report_input=validation["transactions"],
            audit_context=validation["audit_context"],
            input_warnings=validation["warnings"],
        )
        self.assertEqual(report_result["summary"]["inventory_value"], 0.0)
        self.assertEqual(report_result["summary"]["cost_method"], "weighted_average")
        summary_row = report_result["tables"]["summary"][0]
        self.assertEqual(summary_row["Kalan Stok"], -5.0)
        self.assertEqual(summary_row["Toplam Stok Degeri"], 0.0)
        negative_stock_warning = next(w for w in report_result["warnings"] if w["type"] == "negative_stock")
        self.assertEqual(negative_stock_warning["severity"], "warning")
        self.assertEqual(negative_stock_warning["action"], "stock_value_clamped_to_zero")
        self.assertEqual(negative_stock_warning["context"]["remaining_stock"], -5.0)

    def test_inventory_stock_in_movement_valuation_uses_input_unit_cost(self):
        report_definition = self.registry.get_report_definition("inventory_cost_report")
        validation = validate_transactions(sample_inventory_stock_in_records(), report_definition=report_definition)
        report_result = generate_report(
            report_definition=report_definition,
            report_input=validation["transactions"],
            audit_context=validation["audit_context"],
            input_warnings=validation["warnings"],
        )

        movement_row = report_result["tables"]["movements"][0]
        self.assertEqual(movement_row["Urun Kodu"], "STK-002")
        self.assertEqual(movement_row["Degerleme Birim Maliyeti"], 30000.0)
        self.assertEqual(movement_row["Toplam Degerleme Maliyeti"], 240000.0)
        self.assertEqual(movement_row["Maliyet Yontemi"], "Weighted Average")

    def test_inventory_stock_out_without_input_cost_uses_weighted_average(self):
        report_definition = self.registry.get_report_definition("inventory_cost_report")
        validation = validate_transactions(sample_inventory_stock_out_without_cost_records(), report_definition=report_definition)
        report_result = generate_report(
            report_definition=report_definition,
            report_input=validation["transactions"],
            audit_context=validation["audit_context"],
            input_warnings=validation["warnings"],
        )

        stock_out_row = next(row for row in report_result["tables"]["movements"] if row["Islem Tipi"] == "stock_out")
        summary_row = report_result["tables"]["summary"][0]
        self.assertEqual(stock_out_row["Degerleme Birim Maliyeti"], 100.0)
        self.assertEqual(stock_out_row["Toplam Degerleme Maliyeti"], 300.0)
        self.assertEqual(summary_row["Ortalama Birim Maliyet"], 100.0)
        self.assertEqual(summary_row["Toplam Stok Degeri"], 700.0)

    def test_inventory_same_product_code_different_name_merges_and_warns(self):
        report_definition = self.registry.get_report_definition("inventory_cost_report")
        validation = validate_transactions(sample_inventory_inconsistent_name_records(), report_definition=report_definition)
        report_result = generate_report(
            report_definition=report_definition,
            report_input=validation["transactions"],
            audit_context=validation["audit_context"],
            input_warnings=validation["warnings"],
        )

        self.assertEqual(len(report_result["tables"]["summary"]), 1)
        summary_row = report_result["tables"]["summary"][0]
        self.assertEqual(summary_row["Urun Kodu"], "STK-001")
        self.assertEqual(summary_row["Toplam Giris Miktari"], 5.0)
        self.assertEqual(summary_row["Toplam Cikis Miktari"], 2.0)
        inconsistent_warning = next(w for w in report_result["warnings"] if w["type"] == "inconsistent_product_name")
        self.assertEqual(inconsistent_warning["field"], "product_name")
        self.assertEqual(inconsistent_warning["action"], "used_first_display_name")
        self.assertEqual(inconsistent_warning["context"]["product_code"], "STK-001")
        self.assertEqual(sorted(inconsistent_warning["context"]["detected_names"]), ["Laptop", "Yazilim Lisansi"])

    def test_inventory_summary_and_movements_are_consistent_by_inventory_key(self):
        report_definition = self.registry.get_report_definition("inventory_cost_report")
        validation = validate_transactions(sample_inventory_consistency_records(), report_definition=report_definition)
        report_result = generate_report(
            report_definition=report_definition,
            report_input=validation["transactions"],
            audit_context=validation["audit_context"],
            input_warnings=validation["warnings"],
        )

        summary_by_code = {row["Urun Kodu"]: row for row in report_result["tables"]["summary"]}
        movement_rows = report_result["tables"]["movements"]
        for product_code, summary_row in summary_by_code.items():
            product_movements = [row for row in movement_rows if row["Urun Kodu"] == product_code]
            stock_in_qty = sum(row["Miktar"] for row in product_movements if row["Islem Tipi"] == "stock_in")
            stock_out_qty = sum(row["Miktar"] for row in product_movements if row["Islem Tipi"] == "stock_out")
            stock_in_cost = sum(row["Toplam Degerleme Maliyeti"] for row in product_movements if row["Islem Tipi"] == "stock_in")
            expected_avg = round(stock_in_cost / stock_in_qty, 2) if stock_in_qty else 0.0

            self.assertEqual(summary_row["Toplam Giris Miktari"], stock_in_qty)
            self.assertEqual(summary_row["Toplam Cikis Miktari"], stock_out_qty)
            self.assertEqual(summary_row["Ortalama Birim Maliyet"], expected_avg)

    def test_current_account_excludes_paid_rows_from_open_and_aging(self):
        report_definition = self.registry.get_report_definition("current_account_report")
        validation = validate_transactions(sample_current_account_records(), report_definition=report_definition)
        report_result = generate_report(
            report_definition=report_definition,
            report_input=validation["transactions"],
            audit_context=validation["audit_context"],
            input_warnings=validation["warnings"],
        )

        self.assertEqual(report_result["summary"]["open_debt"], 0.0)
        self.assertEqual(report_result["summary"]["open_receivable"], 2000.0)
        aging_row = report_result["tables"]["aging"][0]
        self.assertEqual(aging_row["Borc 0-30 Gun"], 0.0)
        self.assertEqual(aging_row["Alacak 0-30 Gun"], 2000.0)

    def test_debt_receivable_report_risk_score_uses_counterparty_type(self):
        report_definition = self.registry.get_report_definition("debt_receivable_report")
        validation = validate_transactions(sample_debt_records_with_counterparty_type(), report_definition=report_definition)
        report_result = generate_report(
            report_definition=report_definition,
            report_input=validation["transactions"],
            audit_context=validation["audit_context"],
            input_warnings=validation["warnings"],
        )
        main_rows = report_result["tables"]["main"]
        customer_row = next(row for row in main_rows if row["Cari/Firma"] == "ABC Ltd.")
        supplier_row = next(row for row in main_rows if row["Cari/Firma"] == "XYZ Tedarik")
        self.assertEqual(customer_row["Risk Skoru"], 80000.0)
        self.assertEqual(customer_row["Risk Durumu"], "Yuksek")
        self.assertEqual(supplier_row["Risk Skoru"], 10000.0)

    def test_debt_receivable_report_derives_amount_and_direction_from_split_columns(self):
        report_definition = self.registry.get_report_definition("debt_receivable_report")
        validation = validate_transactions(sample_debt_records_with_split_amount_columns(), report_definition=report_definition)

        self.assertTrue(validation["valid"], validation["errors"])
        self.assertEqual(validation["usable_row_count"], 2)
        first_row = validation["transactions"][0]
        second_row = validation["transactions"][1]
        self.assertEqual(float(first_row["amount"]), 50000.0)
        self.assertEqual(first_row["direction"], "debt")
        self.assertEqual(float(second_row["amount"]), 25000.0)
        self.assertEqual(second_row["direction"], "receivable")
        self.assertFalse(any("amount" in (warning.get("field") or "") for warning in validation["warnings"] if warning.get("row") in {1, 2} and warning.get("action") == "row_dropped"))

        report_result = generate_report(
            report_definition=report_definition,
            report_input=validation["transactions"],
            audit_context=validation["audit_context"],
            input_warnings=validation["warnings"],
        )
        main_row = next(row for row in report_result["tables"]["main"] if row["Cari/Firma"] == "ABC Ltd.")
        detail_rows = report_result["tables"]["details"]
        debt_detail = next(row for row in detail_rows if row["Cari/Firma"] == "ABC Ltd.")
        receivable_detail = next(row for row in detail_rows if row["Cari/Firma"] == "XYZ AS")
        self.assertEqual(main_row["Toplam Borc"], 50000.0)
        self.assertEqual(debt_detail["Borc"], 50000.0)
        self.assertEqual(debt_detail["Alacak"], 0.0)
        self.assertEqual(receivable_detail["Borc"], 0.0)
        self.assertEqual(receivable_detail["Alacak"], 25000.0)

    def test_profitability_report_is_future_proof(self):
        report_definition = self.registry.get_report_definition("profitability_report")
        validation = validate_transactions(sample_profitability_records(), report_definition=report_definition)
        report_result = generate_report(
            report_definition=report_definition,
            report_input=validation["transactions"],
            audit_context=validation["audit_context"],
            input_warnings=validation["warnings"],
        )

        self.assertEqual(report_result["summary"]["cash_profit"], 300000.0)
        self.assertIsNone(report_result["summary"]["accounting_profit"])
        self.assertEqual(report_result["summary"]["net_cash_profit_loss"], 300000.0)
        self.assertNotIn("gross_profit", report_result["summary"])

    def test_duplicate_confidence_high_drops_transaction_id_duplicate(self):
        report_definition = self.registry.get_report_definition("vat_summary_report")
        validation = validate_transactions(sample_vat_duplicate_transaction_id_records(), report_definition=report_definition)

        self.assertTrue(validation["valid"], validation["errors"])
        self.assertEqual(len(validation["transactions"]), 1)
        self.assertTrue(any(w["type"] == "duplicate" and w["confidence"] == "high" for w in validation["warnings"]))

    def test_decimal_precision_smoke(self):
        report_definition = self.registry.get_report_definition("sales_performance_report")
        validation = validate_transactions(sample_sales_precision_records(), report_definition=report_definition)
        report_result = generate_report(
            report_definition=report_definition,
            report_input=validation["transactions"],
            audit_context=validation["audit_context"],
            input_warnings=validation["warnings"],
        )
        self.assertEqual(report_result["summary"]["total_sales"], 0.3)

    def test_generate_outputs_returns_metadata_and_execution_contract(self):
        report_definition = self.registry.get_report_definition("personnel_expense_report")
        validation = validate_transactions(sample_personnel_expense_records(), report_definition=report_definition)
        report_result = generate_report(
            report_definition=report_definition,
            report_input=validation["transactions"],
            audit_context=validation["audit_context"],
            input_warnings=validation["warnings"],
        )

        output_dir = None
        try:
            output = generate_outputs(
                report_definition=report_definition,
                intent={"filters": {}},
                transactions=validation["transactions"],
                report_result=report_result,
                source_file=None,
                mapping=None,
                warnings=validation["warnings"],
                audit_context=validation["audit_context"],
            )
            output_dir = Path(output["output_dir"])
            self.assertEqual(output["status"], "success")
            self.assertIn("audit_run_id", output["metadata"])
            self.assertEqual(output["metadata"]["timezone"], settings.DEFAULT_TIMEZONE)
            self.assertEqual(output["metadata"]["calculation_version"], settings.CALCULATION_VERSION)
            self.assertEqual(output["execution_status"], "success")
            self.assertIn("warning_summary", output)
            workbook = load_workbook(Path(output["files"]["xlsx"]), data_only=True)
            self.assertIn("Personel Gider Ozeti", workbook.sheetnames)
        finally:
            if output_dir and output_dir.exists():
                shutil.rmtree(output_dir)

    def test_income_expense_legacy_warning_flow_is_normalized(self):
        report_definition = self.registry.get_report_definition("income_expense_report")
        validation = validate_transactions(sample_income_expense_records_with_invalid_row(), report_definition=report_definition)

        self.assertTrue(validation["valid"], validation["errors"])
        self.assertTrue(validation["warnings"])
        self.assertTrue(all(isinstance(item, dict) for item in validation["warnings"]))

        report_result = generate_report(
            report_definition=report_definition,
            report_input=validation["transactions"],
            audit_context=validation["audit_context"],
            input_warnings=validation["warnings"],
        )
        self.assertTrue(all(isinstance(item, dict) for item in report_result["warnings"]))

        output_dir = None
        try:
            output = generate_outputs(
                report_definition=report_definition,
                intent={"filters": {}},
                transactions=validation["transactions"],
                report_result=report_result,
                source_file=None,
                mapping=None,
                warnings=validation["warnings"],
                audit_context=validation["audit_context"],
            )
            output_dir = Path(output["output_dir"])
            self.assertTrue(all(isinstance(item, dict) for item in output["warnings"]))
            self.assertIn("warning_summary", output)
        finally:
            if output_dir and output_dir.exists():
                shutil.rmtree(output_dir)

    def test_sales_report_warning_schema_contains_audit_and_lineage(self):
        report_definition = self.registry.get_report_definition("sales_performance_report")
        validation = validate_transactions(sample_sales_records_with_refund(), report_definition=report_definition)

        recalculated_warning = next(w for w in validation["warnings"] if w["type"] == "recalculated_total_sales" and w["field"] == "total_sales")
        self.assertEqual(recalculated_warning["audit_run_id"], validation["audit_context"]["audit_run_id"])
        self.assertEqual(recalculated_warning["calculated_from"], ["quantity", "unit_price", "discount"])
        self.assertEqual(recalculated_warning["lineage"]["rule"], "signed(quantity * unit_price - discount)")
        self.assertEqual(recalculated_warning["action"], "used_calculated_value")

    def test_timezone_normalization_uses_canonical_istanbul_period(self):
        report_definition = self.registry.get_report_definition("tax_calculation_report")
        validation = validate_transactions(sample_timezone_boundary_tax_records(), report_definition=report_definition)

        self.assertTrue(validation["valid"], validation["errors"])
        self.assertEqual(validation["transactions"][0]["date"], "2026-05-01T01:30:00+03:00")

        report_result = generate_report(
            report_definition=report_definition,
            report_input=validation["transactions"],
            audit_context=validation["audit_context"],
            input_warnings=validation["warnings"],
        )
        periodic_rows = report_result["tables"]["periodic"]
        self.assertEqual(periodic_rows[0]["Donem"], "2026-05")

    def test_mixed_currency_summary_warns_and_partitions(self):
        report_definition = self.registry.get_report_definition("sales_performance_report")
        validation = validate_transactions(sample_sales_mixed_currency_records(), report_definition=report_definition)
        report_result = generate_report(
            report_definition=report_definition,
            report_input=validation["transactions"],
            audit_context=validation["audit_context"],
            input_warnings=validation["warnings"],
        )

        self.assertIsNone(report_result["summary"]["total_sales"])
        self.assertTrue(report_result["summary"]["mixed_currency_detected"])
        self.assertEqual(report_result["summary"]["currencies_detected"], ["TRY", "USD"])
        self.assertIn("TRY", report_result["summary"]["totals_by_currency"])
        self.assertIn("USD", report_result["summary"]["totals_by_currency"])
        self.assertTrue(any(w["type"] == "mixed_currency" and w["severity"] == "critical" for w in report_result["warnings"]))

    def test_masked_export_adds_warning_and_hides_sensitive_values(self):
        report_definition = self.registry.get_report_definition("personnel_expense_report")
        validation = validate_transactions(sample_personnel_expense_records(), report_definition=report_definition)
        report_result = generate_report(
            report_definition=report_definition,
            report_input=validation["transactions"],
            audit_context=validation["audit_context"],
            input_warnings=validation["warnings"],
        )

        output_dir = None
        original_mask_setting = settings.MASK_SENSITIVE_EXPORTS
        try:
            object.__setattr__(settings, "MASK_SENSITIVE_EXPORTS", True)
            output = generate_outputs(
                report_definition=report_definition,
                intent={"filters": {}},
                transactions=validation["transactions"],
                report_result=report_result,
                source_file=None,
                mapping=None,
                warnings=validation["warnings"],
                audit_context=validation["audit_context"],
            )
            output_dir = Path(output["output_dir"])
            payload = json.loads(Path(output["files"]["json"]).read_text(encoding="utf-8"))
            self.assertTrue(any(w["type"] == "sensitive_data_masked" for w in payload["warnings"]))
            self.assertEqual(payload["transactions"][0]["employee_name"], "A**** Y*****")
        finally:
            object.__setattr__(settings, "MASK_SENSITIVE_EXPORTS", original_mask_setting)
            if output_dir and output_dir.exists():
                shutil.rmtree(output_dir)

    def test_prompt_parser_accepts_json_and_alias_keys(self):
        payload = json.dumps(
            [
                {
                    "Tarih": "2026-05-01",
                    "Musteri": "ABC Ltd.",
                    "Urun Adi": "Danismanlik Hizmeti",
                    "Miktar": 2,
                    "Birim Fiyat": 50000,
                    "Indirim": 5000,
                }
            ],
            ensure_ascii=False,
        )
        extraction = extract_prompt_transactions(
            prompt=payload,
            report_type="sales_performance_report",
            intent={"report_type": "sales_performance_report"},
        )

        self.assertEqual(extraction["status"], "passed")
        self.assertEqual(extraction["transactions"][0]["customer"], "ABC Ltd.")
        self.assertEqual(extraction["transactions"][0]["product_name"], "Danismanlik Hizmeti")

    def test_report_filters_leave_data_untouched_when_prompt_empty(self):
        df = pd.DataFrame(sample_debt_records_with_counterparty_type())
        filtered_df, warnings, summary = apply_report_filters(
            df,
            filter_spec=ReportFilterSpec(),
            report_type="debt_receivable_report",
        )

        self.assertEqual(len(filtered_df), len(df))
        self.assertFalse(summary.applied)
        self.assertEqual(summary.filtered_row_count, len(df))
        self.assertEqual(warnings, [])

    @patch("services.report_prompt_filter_service.today_date", return_value=date(2026, 5, 18))
    def test_report_filter_parser_extracts_last_three_months(self, _mock_today):
        spec = extract_report_filters(
            user_prompt="Son 3 ayın verilerini kullanarak hazırla.",
            report_type="sales_performance_report",
            available_columns=["date", "total_sales"],
            normalized_schema=["date", "total_sales"],
        )

        self.assertIsNotNone(spec.date_range)
        self.assertEqual(spec.date_range.field, "date")
        self.assertEqual(spec.date_range.start_date, "2026-02-18")
        self.assertEqual(spec.date_range.end_date, "2026-05-18")

    @patch("services.report_prompt_filter_service.today_date", return_value=date(2026, 5, 19))
    def test_report_filter_parser_extracts_month_boundary_range(self, _mock_today):
        spec = extract_report_filters(
            user_prompt="mart ayının başından itibaren nisan ayının sonuna kadar olan zaman aralığında rapor hazırla",
            report_type="income_expense_report",
            available_columns=["date", "amount", "description", "direction"],
            normalized_schema=["date", "amount", "description", "direction"],
        )

        self.assertIsNotNone(spec.date_range)
        self.assertEqual(spec.date_range.start_date, "2026-03-01")
        self.assertEqual(spec.date_range.end_date, "2026-04-30")

    def test_coerce_error_items_wraps_string_errors_for_api_model(self):
        errors = coerce_error_items(["Filtre ve validation sonrasi rapor uretecek veri kalmadi."])

        self.assertEqual(
            errors,
            [{"message": "Filtre ve validation sonrasi rapor uretecek veri kalmadi."}],
        )

    def test_date_in_range_accepts_iso_timestamps(self):
        self.assertTrue(date_in_range("2026-03-15T10:30:00+03:00", start="2026-03-01", end="2026-04-30"))
        self.assertFalse(date_in_range("2026-05-01T00:00:00+03:00", start="2026-03-01", end="2026-04-30"))

    def test_report_filter_engine_applies_amount_threshold(self):
        records = sample_debt_records_with_counterparty_type()
        spec = extract_report_filters(
            user_prompt="Sadece 50 bin üzeri işlemleri dahil et.",
            report_type="debt_receivable_report",
            available_columns=list(records[0].keys()),
            normalized_schema=list(records[0].keys()),
        )

        filtered_df, warnings, summary = apply_report_filters(
            pd.DataFrame(records),
            filter_spec=spec,
            report_type="debt_receivable_report",
            user_prompt="Sadece 50 bin üzeri işlemleri dahil et.",
        )

        self.assertEqual(len(filtered_df), 1)
        self.assertEqual(filtered_df.iloc[0]["counterparty"], "ABC Ltd.")
        self.assertTrue(summary.applied)
        self.assertEqual(summary.filtered_row_count, 1)
        self.assertEqual(warnings, [])

    @patch("services.report_prompt_filter_service.today_date", return_value=date(2026, 5, 18))
    def test_report_filter_engine_combines_date_and_amount_filters(self, _mock_today):
        records = sample_debt_filter_records()
        spec = extract_report_filters(
            user_prompt="son 3 ay ve 50 bin üzeri işlemler",
            report_type="debt_receivable_report",
            available_columns=list(records[0].keys()),
            normalized_schema=list(records[0].keys()),
        )

        filtered_df, warnings, summary = apply_report_filters(
            pd.DataFrame(records),
            filter_spec=spec,
            report_type="debt_receivable_report",
            user_prompt="son 3 ay ve 50 bin üzeri işlemler",
        )

        self.assertEqual(len(filtered_df), 1)
        self.assertEqual(filtered_df.iloc[0]["counterparty"], "Beta Ltd.")
        self.assertEqual(summary.filtered_row_count, 1)
        self.assertTrue(any("2026-02-18" in line for line in summary.summary_lines))
        self.assertEqual(warnings, [])

    def test_report_filter_engine_keeps_only_unpaid_rows(self):
        records = sample_current_account_records()
        spec = extract_report_filters(
            user_prompt="Sadece ödenmemişler",
            report_type="current_account_report",
            available_columns=list(records[0].keys()),
            normalized_schema=list(records[0].keys()),
        )

        filtered_df, warnings, summary = apply_report_filters(
            pd.DataFrame(records),
            filter_spec=spec,
            report_type="current_account_report",
            user_prompt="Sadece ödenmemişler",
        )

        self.assertEqual(len(filtered_df), 1)
        self.assertEqual(filtered_df.iloc[0]["payment_status"], "unpaid")
        self.assertEqual(summary.filtered_row_count, 1)
        self.assertEqual(warnings, [])

    def test_report_filter_engine_applies_department_filter_case_insensitively(self):
        records = sample_personnel_filter_records()
        spec = extract_report_filters(
            user_prompt="Personel gider raporunda sadece yazılım departmanını kullan.",
            report_type="personnel_expense_report",
            available_columns=list(records[0].keys()),
            normalized_schema=list(records[0].keys()),
        )

        filtered_df, warnings, summary = apply_report_filters(
            pd.DataFrame(records),
            filter_spec=spec,
            report_type="personnel_expense_report",
            user_prompt="Personel gider raporunda sadece yazılım departmanını kullan.",
        )

        self.assertEqual(len(filtered_df), 2)
        self.assertEqual(set(filtered_df["department"]), {"Yazılım"})
        self.assertEqual(summary.filtered_row_count, 2)
        self.assertEqual(warnings, [])

    def test_report_filter_engine_warns_when_field_is_missing(self):
        records = sample_personnel_filter_records()
        spec = ReportFilterSpec(
            category_filters=[CategoryFilter(field="region", values=["İstanbul"])],
            confidence=1.0,
        )

        filtered_df, warnings, summary = apply_report_filters(
            pd.DataFrame(records),
            filter_spec=spec,
            report_type="personnel_expense_report",
            user_prompt="Sadece İstanbul bölgesini kullan.",
        )

        self.assertEqual(len(filtered_df), len(records))
        self.assertEqual(summary.filtered_row_count, len(records))
        self.assertTrue(any(w["type"] == "filter_field_missing" for w in warnings))

    def test_report_filter_engine_returns_no_rows_warning(self):
        records = sample_personnel_filter_records()
        spec = ReportFilterSpec(
            category_filters=[CategoryFilter(field="department", values=["Destek"])],
            confidence=1.0,
        )

        filtered_df, warnings, summary = apply_report_filters(
            pd.DataFrame(records),
            filter_spec=spec,
            report_type="personnel_expense_report",
            user_prompt="Sadece Destek departmanı",
        )

        self.assertTrue(filtered_df.empty)
        self.assertEqual(summary.filtered_row_count, 0)
        self.assertTrue(any(w["type"] == "filter_no_rows_remaining" for w in warnings))

    def test_report_filter_engine_applies_group_top_n_customer_filter(self):
        records = sample_sales_customer_ranking_records()
        spec = extract_report_filters(
            user_prompt="en büyük 5 müşteri",
            report_type="sales_performance_report",
            available_columns=list(records[0].keys()),
            normalized_schema=list(records[0].keys()),
        )

        filtered_df, warnings, summary = apply_report_filters(
            pd.DataFrame(records),
            filter_spec=spec,
            report_type="sales_performance_report",
            user_prompt="en büyük 5 müşteri",
        )

        self.assertEqual(sorted(filtered_df["customer"].unique().tolist()), ["A Müşteri", "B Müşteri", "C Müşteri", "D Müşteri", "E Müşteri"])
        self.assertNotIn("F Müşteri", filtered_df["customer"].unique().tolist())
        self.assertTrue(any("Top 5 customer" == line for line in summary.summary_lines))
        self.assertEqual(warnings, [])

    def test_report_filter_engine_applies_risk_based_counterparty_ranking(self):
        records = sample_current_account_risk_records()
        spec = extract_report_filters(
            user_prompt="en riskli cariler",
            report_type="current_account_report",
            available_columns=list(records[0].keys()),
            normalized_schema=list(records[0].keys()),
        )
        self.assertIsNotNone(spec.ranking)
        self.assertEqual(spec.ranking.metric_field, "risk_score")
        spec.ranking.top_n = 1

        filtered_df, warnings, summary = apply_report_filters(
            pd.DataFrame(records),
            filter_spec=spec,
            report_type="current_account_report",
            user_prompt="en riskli cariler",
        )

        self.assertEqual(filtered_df["counterparty"].unique().tolist(), ["Riskli A.Ş."])
        self.assertEqual(summary.filtered_row_count, 1)
        self.assertEqual(warnings, [])

    def test_generate_outputs_persists_filter_metadata_and_filter_sheet(self):
        report_definition = self.registry.get_report_definition("sales_performance_report")
        validation = validate_transactions(sample_sales_customer_ranking_records(), report_definition=report_definition)
        report_result = generate_report(
            report_definition=report_definition,
            report_input=validation["transactions"],
            audit_context=validation["audit_context"],
            input_warnings=validation["warnings"],
        )

        spec = extract_report_filters(
            user_prompt="en büyük 5 müşteri",
            report_type="sales_performance_report",
            available_columns=list(validation["transactions"][0].keys()),
            normalized_schema=list(validation["transactions"][0].keys()),
        )
        filtered_df, _, filter_summary = apply_report_filters(
            pd.DataFrame(validation["transactions"]),
            filter_spec=spec,
            report_type="sales_performance_report",
            user_prompt="en büyük 5 müşteri",
        )
        report_result["filter"] = filter_summary.model_dump()
        report_result.setdefault("metadata", {})
        report_result["metadata"]["filter"] = filter_summary.model_dump()
        report_result = append_filter_sheet(report_result, filter_summary)

        output_dir = None
        try:
            output = generate_outputs(
                report_definition=report_definition,
                intent={"filters": {}},
                transactions=filtered_df.to_dict(orient="records"),
                report_result=report_result,
                source_file=None,
                mapping=None,
                warnings=validation["warnings"],
                audit_context=validation["audit_context"],
            )
            output_dir = Path(output["output_dir"])
            payload = json.loads(Path(output["files"]["json"]).read_text(encoding="utf-8"))
            workbook = load_workbook(Path(output["files"]["xlsx"]), data_only=True)

            self.assertIn("Rapor Filtreleri", workbook.sheetnames)
            self.assertEqual(payload["filter"]["filtered_row_count"], 5)
            self.assertEqual(payload["metadata"]["filter"]["input_row_count"], 6)
        finally:
            if output_dir and output_dir.exists():
                shutil.rmtree(output_dir)


def sample_personnel_expense_records():
    return [
        {
            "date": "2026-05-31",
            "employee_name": "Ahmet Yilmaz",
            "department": "Satis",
            "gross_salary": 37000,
            "bonus": 2000,
            "benefits": 4000,
        },
        {
            "date": "2026-05-31",
            "employee_name": "Ayse Demir",
            "department": "Satis",
            "gross_salary": 30000,
            "bonus": 1000,
            "benefits": 2000,
            "employer_extra_cost": 500,
            "employer_cost": 6750,
        },
    ]


def sample_payroll_records():
    return [
        {
            "date": "2026-05-31",
            "employee_name": "A",
            "gross_salary": 60000,
            "net_salary": 50000,
            "bonus": 4000,
            "benefits": 0,
        }
    ]


def sample_personnel_expense_record_with_input_total_cost():
    return [
        {
            "date": "2026-05-31",
            "employee_name": "Ahmet Yilmaz",
            "department": "Satis",
            "gross_salary": 37000,
            "bonus": 2000,
            "benefits": 4000,
            "employer_cost": 8325,
            "total_employer_cost": 51325,
        }
    ]


def sample_vat_records_decimal_rate():
    return [
        {
            "date": "2026-05-01",
            "invoice_no": "A123",
            "counterparty": "XYZ AS",
            "transaction_type": "sale",
            "base_amount": 100000,
            "tax_rate": 0.20,
        },
        {
            "date": "2026-05-02",
            "invoice_no": "B100",
            "counterparty": "Tedarikci",
            "transaction_type": "purchase",
            "base_amount": 20000,
            "tax_rate": 20,
        },
    ]


def sample_invalid_tax_rate_records():
    return [
        {
            "date": "2026-05-01",
            "tax_type": "KDV",
            "base_amount": 100000,
            "tax_rate": 120,
            "transaction_type": "sale",
        }
    ]


def sample_tax_records_decimal_rate():
    return [
        {
            "date": "2026-05-01",
            "tax_type": "KDV",
            "base_amount": 100000,
            "tax_rate": 0.20,
            "transaction_type": "sale",
        },
        {
            "date": "2026-05-10",
            "tax_type": "Stopaj",
            "base_amount": 40000,
            "tax_rate": 10,
            "transaction_type": "expense",
        },
    ]


def sample_sales_records_with_refund():
    return [
        {
            "date": "2026-05-01",
            "customer": "ABC Ltd.",
            "product_name": "Danismanlik Hizmeti",
            "quantity": 2,
            "unit_price": 50000,
            "discount": 5000,
            "total_sales": 10,
            "salesperson": "Ayse",
        },
        {
            "date": "2026-05-03",
            "customer": "ABC Ltd.",
            "product_name": "Danismanlik Hizmeti",
            "transaction_type": "refund",
            "quantity": 1,
            "total_sales": -50000,
            "salesperson": "Ayse",
        },
    ]


def sample_sales_records_with_unknown_type():
    return [
        {
            "date": "2026-05-01",
            "customer": "ABC Ltd.",
            "product_name": "Urun A",
            "quantity": 1,
            "unit_price": 1000,
            "transaction_type": "sale",
        },
        {
            "date": "2026-05-02",
            "customer": "XYZ Ltd.",
            "product_name": "Urun B",
            "quantity": 1,
            "unit_price": 500,
            "transaction_type": "weird",
        },
    ]


def sample_sales_records_with_return_status_only():
    return [
        {
            "date": "2026-05-01",
            "customer": "ABC Ltd.",
            "product_name": "Urun A",
            "quantity": 2,
            "unit_price": 500,
            "return_status": "partial_return",
        },
        {
            "date": "2026-05-02",
            "customer": "ABC Ltd.",
            "product_name": "Urun B",
            "quantity": 1,
            "unit_price": 1000,
            "return_status": "none",
        },
    ]


def sample_sales_excel_rows_with_return_status_column():
    return [
        {
            "Tarih": "2026-05-01",
            "Musteri": "ABC Ltd.",
            "Urun": "Urun A",
            "Miktar": 2,
            "Birim Fiyat": 500,
            "Return Status": "partial_return",
        },
        {
            "Tarih": "2026-05-02",
            "Musteri": "ABC Ltd.",
            "Urun": "Urun B",
            "Miktar": 1,
            "Birim Fiyat": 1000,
            "Return Status": "none",
        },
    ]


def sample_negative_sale_records():
    return [
        {
            "date": "2026-05-01",
            "customer": "ABC Ltd.",
            "product_name": "Urun A",
            "transaction_type": "sale",
            "quantity": 1,
            "total_sales": -1000,
        }
    ]


def sample_sales_records_all_unknown_type():
    return [
        {
            "date": "2026-05-01",
            "customer": "ABC Ltd.",
            "product_name": "Urun A",
            "quantity": 1,
            "unit_price": 1000,
            "transaction_type": "weird_a",
        },
        {
            "date": "2026-05-02",
            "customer": "XYZ Ltd.",
            "product_name": "Urun B",
            "quantity": 2,
            "unit_price": 1500,
            "transaction_type": "weird_b",
        },
    ]


def sample_sales_leaderboard_records():
    return [
        {
            "date": "2026-05-01",
            "customer": "Alpha AS",
            "product_name": "Urun A",
            "quantity": 3,
            "unit_price": 1000,
            "transaction_type": "sale",
            "salesperson": "Deniz",
        },
        {
            "date": "2026-05-02",
            "customer": "Alpha AS",
            "product_name": "Urun A",
            "quantity": 2,
            "total_sales": -2000,
            "transaction_type": "refund",
            "salesperson": "Deniz",
        },
        {
            "date": "2026-05-03",
            "customer": "Beta AS",
            "product_name": "Urun B",
            "quantity": 2,
            "unit_price": 1200,
            "transaction_type": "sale",
            "salesperson": "Selin",
        },
    ]


def sample_inventory_negative_stock_records():
    return [
        {"date": "2026-05-01", "product_code": "STK-NEG", "product_name": "Laptop", "quantity": 5, "unit_cost": 1000, "transaction_type": "stock_in"},
        {"date": "2026-05-02", "product_code": "STK-NEG", "product_name": "Laptop", "quantity": 10, "transaction_type": "stock_out"},
    ]


def sample_inventory_stock_in_records():
    return [
        {
            "date": "2026-05-01",
            "product_code": "STK-002",
            "product_name": "Danismanlik Hizmeti",
            "quantity": 8,
            "unit_cost": 30000,
            "transaction_type": "stock_in",
        }
    ]


def sample_inventory_stock_out_without_cost_records():
    return [
        {
            "date": "2026-05-01",
            "product_code": "STK-100",
            "product_name": "Klavye",
            "quantity": 10,
            "unit_cost": 100,
            "transaction_type": "stock_in",
        },
        {
            "date": "2026-05-03",
            "product_code": "STK-100",
            "product_name": "Klavye",
            "quantity": 3,
            "unit_cost": None,
            "transaction_type": "stock_out",
        },
    ]


def sample_inventory_inconsistent_name_records():
    return [
        {
            "date": "2026-05-01",
            "product_code": "STK-001",
            "product_name": "Laptop",
            "quantity": 5,
            "unit_cost": 100,
            "transaction_type": "stock_in",
        },
        {
            "date": "2026-05-02",
            "product_code": "STK-001",
            "product_name": "Yazilim Lisansi",
            "quantity": 2,
            "transaction_type": "stock_out",
        },
    ]


def sample_inventory_consistency_records():
    return [
        {
            "date": "2026-05-01",
            "product_code": "STK-011",
            "product_name": "Laptop",
            "quantity": 20,
            "unit_cost": 30000,
            "transaction_type": "stock_in",
        },
        {
            "date": "2026-05-05",
            "product_code": "STK-011",
            "product_name": "Laptop",
            "quantity": 5,
            "transaction_type": "stock_out",
        },
        {
            "date": "2026-05-02",
            "product_code": "STK-004",
            "product_name": "Monitor",
            "quantity": 9,
            "unit_cost": 37000,
            "transaction_type": "stock_in",
        },
    ]


def sample_current_account_records():
    return [
        {
            "date": "2026-05-01",
            "counterparty": "ABC",
            "amount": 1000,
            "transaction_direction": "debt",
            "due_date": "2026-05-05",
            "payment_status": "paid",
        },
        {
            "date": "2026-05-02",
            "counterparty": "ABC",
            "amount": 2000,
            "transaction_direction": "receivable",
            "due_date": "2026-05-10",
            "payment_status": "unpaid",
        },
    ]


def sample_debt_filter_records():
    return [
        {
            "date": "2026-01-10",
            "counterparty": "Alpha Ltd.",
            "amount": 80000,
            "direction": "receivable",
            "counterparty_type": "customer",
            "due_date": "2026-01-25",
            "payment_status": "unpaid",
        },
        {
            "date": "2026-03-20",
            "counterparty": "Beta Ltd.",
            "amount": 60000,
            "direction": "receivable",
            "counterparty_type": "customer",
            "due_date": "2026-03-30",
            "payment_status": "unpaid",
        },
        {
            "date": "2026-04-25",
            "counterparty": "Gamma Ltd.",
            "amount": 40000,
            "direction": "receivable",
            "counterparty_type": "customer",
            "due_date": "2026-05-10",
            "payment_status": "unpaid",
        },
    ]


def sample_personnel_filter_records():
    return [
        {
            "date": "2026-05-31",
            "employee_name": "Ada",
            "department": "Yazılım",
            "gross_salary": 60000,
            "bonus": 5000,
            "benefits": 3000,
            "total_employer_cost": 82000,
        },
        {
            "date": "2026-05-31",
            "employee_name": "Ece",
            "department": "Finans",
            "gross_salary": 45000,
            "bonus": 2000,
            "benefits": 1500,
            "total_employer_cost": 61000,
        },
        {
            "date": "2026-05-31",
            "employee_name": "Mert",
            "department": "Yazılım",
            "gross_salary": 55000,
            "bonus": 4000,
            "benefits": 2500,
            "total_employer_cost": 76000,
        },
    ]


def sample_debt_records_with_counterparty_type():
    return [
        {
            "date": "2026-05-01",
            "counterparty": "ABC Ltd.",
            "amount": 80000,
            "direction": "receivable",
            "counterparty_type": "customer",
            "due_date": "2026-05-10",
            "payment_status": "unpaid",
        },
        {
            "date": "2026-05-01",
            "counterparty": "XYZ Tedarik",
            "amount": 10000,
            "direction": "debt",
            "counterparty_type": "supplier",
            "due_date": "2026-05-10",
            "payment_status": "unpaid",
        },
    ]


def sample_debt_records_with_split_amount_columns():
    return [
        {
            "date": "2026-05-01",
            "counterparty": "ABC Ltd.",
            "debt_amount": 50000,
            "receivable_amount": 0,
            "due_date": "2026-05-10",
            "payment_status": "unpaid",
        },
        {
            "date": "2026-05-02",
            "counterparty": "XYZ AS",
            "debt_amount": 0,
            "receivable_amount": 25000,
            "due_date": "2026-05-15",
            "payment_status": "unpaid",
        },
        {
            "date": "2026-05-03",
            "counterparty": "Bos Satir Ltd.",
            "debt_amount": 0,
            "receivable_amount": 0,
            "due_date": "2026-05-20",
            "payment_status": "unpaid",
        },
    ]


def sample_profitability_records():
    return [
        {"date": "2026-05-01", "amount": 1000000, "direction": "income", "category": "Satis Geliri"},
        {"date": "2026-05-05", "amount": 700000, "direction": "expense", "category": "Operasyon"},
    ]


def sample_sales_customer_ranking_records():
    return [
        {"date": "2026-05-01", "customer": "A Müşteri", "product_name": "Ürün A", "salesperson": "Selin", "quantity": 1, "unit_price": 120000, "transaction_type": "sale", "total_sales": 120000},
        {"date": "2026-05-01", "customer": "B Müşteri", "product_name": "Ürün B", "salesperson": "Deniz", "quantity": 1, "unit_price": 95000, "transaction_type": "sale", "total_sales": 95000},
        {"date": "2026-05-01", "customer": "C Müşteri", "product_name": "Ürün C", "salesperson": "Mert", "quantity": 1, "unit_price": 80000, "transaction_type": "sale", "total_sales": 80000},
        {"date": "2026-05-01", "customer": "D Müşteri", "product_name": "Ürün D", "salesperson": "Ece", "quantity": 1, "unit_price": 65000, "transaction_type": "sale", "total_sales": 65000},
        {"date": "2026-05-01", "customer": "E Müşteri", "product_name": "Ürün E", "salesperson": "Ada", "quantity": 1, "unit_price": 50000, "transaction_type": "sale", "total_sales": 50000},
        {"date": "2026-05-01", "customer": "F Müşteri", "product_name": "Ürün F", "salesperson": "Can", "quantity": 1, "unit_price": 15000, "transaction_type": "sale", "total_sales": 15000},
    ]


def sample_current_account_risk_records():
    return [
        {
            "date": "2026-05-01",
            "counterparty": "Riskli A.Ş.",
            "counterparty_type": "customer",
            "amount": 120000,
            "transaction_direction": "receivable",
            "due_date": "2026-04-10",
            "payment_status": "unpaid",
        },
        {
            "date": "2026-05-03",
            "counterparty": "Düşük Risk Ltd.",
            "counterparty_type": "customer",
            "amount": 50000,
            "transaction_direction": "receivable",
            "due_date": "2026-06-10",
            "payment_status": "unpaid",
        },
        {
            "date": "2026-05-05",
            "counterparty": "Kapalı Tedarik",
            "counterparty_type": "supplier",
            "amount": 70000,
            "transaction_direction": "debt",
            "due_date": "2026-04-05",
            "payment_status": "paid",
        },
    ]


def sample_vat_duplicate_transaction_id_records():
    return [
        {
            "date": "2026-05-01",
            "invoice_no": "A123",
            "transaction_id": "tx-1",
            "counterparty": "XYZ AS",
            "transaction_type": "sale",
            "base_amount": 100000,
            "tax_rate": 20,
        },
        {
            "date": "2026-05-01",
            "invoice_no": "A124",
            "transaction_id": "tx-1",
            "counterparty": "XYZ AS",
            "transaction_type": "sale",
            "base_amount": 100000,
            "tax_rate": 20,
        },
    ]


def sample_sales_precision_records():
    return [
        {
            "date": "2026-05-01",
            "customer": "Mini Co",
            "product_name": "Precision Item",
            "quantity": 3,
            "unit_price": 0.1,
            "discount": 0,
        }
    ]


def sample_income_expense_records_with_invalid_row():
    return [
        {
            "date": "2026-05-01",
            "amount": 1000,
            "description": "Danismanlik Geliri",
            "direction": "income",
            "category": "Satis",
        },
        {
            "date": "2026-05-02",
            "amount": -500,
            "description": "Gecersiz Satir",
            "direction": "expense",
            "category": "Operasyon",
        },
    ]


def sample_timezone_boundary_tax_records():
    return [
        {
            "date": "2026-04-30T22:30:00Z",
            "tax_type": "KDV",
            "base_amount": 1000,
            "tax_rate": 20,
            "transaction_type": "sale",
            "timezone": "UTC",
        }
    ]


def sample_sales_mixed_currency_records():
    return [
        {
            "date": "2026-05-01",
            "customer": "ABC Ltd.",
            "product_name": "Danismanlik Hizmeti",
            "quantity": 1,
            "unit_price": 1000,
            "currency": "TRY",
        },
        {
            "date": "2026-05-02",
            "customer": "XYZ Inc.",
            "product_name": "Consulting",
            "quantity": 1,
            "unit_price": 100,
            "currency": "USD",
        },
    ]


if __name__ == "__main__":
    unittest.main()
